#!/usr/bin/env python3

from __future__ import annotations

import copy
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
UK_CODEBOOK = ROOT / "documentation" / "SimPaths_Variable_CodebookUK.xlsx"
OUTPUT_CODEBOOK = ROOT / "documentation" / "SimPathsEU_variable_Codebook.xlsx"
TEMP_OUTPUT_CODEBOOK = ROOT / "documentation" / "SimPathsEU_variable_Codebook.__tmp__.xlsx"


CORE_JAVA_FILES = [
    ROOT / "src/main/java/simpaths/model/Person.java",
    ROOT / "src/main/java/simpaths/model/BenefitUnit.java",
    ROOT / "src/main/java/simpaths/model/Household.java",
    ROOT / "src/main/java/simpaths/data/statistics/Statistics.java",
    ROOT / "src/main/java/simpaths/data/statistics/Statistics2.java",
    ROOT / "src/main/java/simpaths/data/statistics/AlignmentAdjustmentFactors.java",
]

METADATA_JAVA_FILES = [
    ROOT / "src/main/java/simpaths/data/startingpop/Processed.java",
]

CONTROL_ENUM_FILES = {
    ROOT / "src/main/java/simpaths/model/enums/TimeSeriesVariable.java": "TimeSeriesVariable",
    ROOT / "src/main/java/simpaths/model/enums/AlignmentVariable.java": "AlignmentVariable",
}

DEFERRED_ENUMS = {
    ROOT / "src/main/java/simpaths/model/Person.java": ["IntegerVariables", "DoublesVariables"],
    ROOT / "src/main/java/simpaths/model/Validator.java": ["DoublesVariables"],
    ROOT / "src/main/java/simpaths/experiment/SimPathsObserver.java": ["LongVariables"],
}


COMMENT_RE = re.compile(r"/\*.*?\*/|//.*?$", re.S | re.M)
FIELD_RE = re.compile(
    r"^(?P<prefix>\s*(?:@[A-Za-z_][\w.()=\" ,]+\s*)*\b(?:public|protected|private)\s+(?:static\s+)?(?:final\s+)?)"
    r"(?P<type>[^;=(){}]+?)\s+"
    r"(?P<name>[A-Za-z_][A-Za-z0-9_]*)\s*(?:=[^;]*)?;\s*$",
    re.M,
)
COLUMN_RE = re.compile(r'@Column\s*\(\s*name\s*=\s*"([^"]+)"')
ENUM_DECL_RE = re.compile(r"\benum\s+([A-Za-z_][A-Za-z0-9_]*)\s*\{", re.M)


HEADER_STYLE_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
MATCHED_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
ADAPTED_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
NEW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
UK_ONLY_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
DEFERRED_FILL = PatternFill(fill_type="solid", fgColor="EDEDED")
REVIEW_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")

MODULE_ORDER = {
    "dem": 1,
    "edu": 2,
    "health": 3,
    "care": 4,
    "lab": 5,
    "y": 6,
    "x": 7,
    "stat": 8,
    "covid": 9,
    "wealth": 10,
    "id": 11,
    "wgt": 12,
}


EXCLUDED_REASONS = {
    "log": "Logger infrastructure",
    "model": "Simulation manager reference",
    "collector": "Collector/observer reference",
    "key": "Persistence key",
    "benefitUnit": "Entity association",
    "household": "Entity association",
    "members": "Entity association",
    "benefitUnits": "Entity association",
    "processed": "Metadata association",
    "personIdCounter": "Static ID counter",
    "benefitUnitIdCounter": "Static ID counter",
    "householdIdCounter": "Static ID counter",
    "ioFlag": "Regression helper flag",
    "states": "Decision-state helper object",
    "taxDbMatch": "Tax matching helper object",
    "sIndexYearMap": "Derived series cache",
    "personContinuousHoursLabourSupplyMap": "Derived hours cache",
    "countMale": "Temporary diagnostic accumulator",
    "countFemale": "Temporary diagnostic accumulator",
    "reportedMissing": "Missing-value reporting helper",
    "lastYear": "Cache/incremental helper",
    "cachedMaleAtRiskOfWork": "Cache field",
    "cachedFemaleAtRiskOfWork": "Cache field",
    "labourChoiceCacheYear": "Cache field",
    "labourChoiceCacheKey": "Cache field",
    "cachedPossibleLabourCombinations": "Cache field",
    "cachedEvalByLabourPairs": "Cache field",
    "cachedUtilityScoreByLabourPairs": "Cache field",
    "labourScoreCacheYear": "Cache field",
    "labourScoreCacheKey": "Cache field",
    "households": "Metadata collection",
    "benefitUnits": "Metadata collection",
    "persons": "Metadata collection",
}


@dataclass
class Candidate:
    source_name: str
    source_kind: str
    java_class: str
    java_path: Path
    order: int
    sibling_name: str | None = None


@dataclass
class ManualSpec:
    status: str
    source_kind: str
    variable_name: str | None = None
    reference_old: str | None = None
    reference_new: str | None = None
    match_basis: str = "Manual review"
    review_required: str = "No"
    review_note: str = ""
    overrides: dict[str, object] | None = None


MANUAL_SPECS = {
    "idMotherImmutable": ManualSpec(
        status="New vs UK",
        source_kind="Java field",
        variable_name="idMotherImmutable",
        overrides={
            "Local": None,
            "Mod": "id",
            "Main attribute": "Mother",
            "Level": None,
            "Measurement": None,
            "Other info": "Immutable",
            "Statistics type": None,
            "Periodicity": None,
            "Description": "Immutable mother ID used to preserve the original maternal link in simulation.",
            "Initial population": "No",
            "Module": "ID",
            "Notes": "No direct UK equivalent found in the reviewed workbook.",
        },
    ),
    "immutable_mother_id": ManualSpec(
        status="New vs UK",
        source_kind="DB column",
        variable_name="idMotherImmutable",
        overrides={
            "Local": None,
            "Mod": "id",
            "Main attribute": "Mother",
            "Other info": "Immutable",
            "Statistics type": None,
            "Description": "Database column backing the immutable mother ID.",
            "Initial population": "No",
            "Module": "ID",
            "Notes": "Database column paired with idMotherImmutable.",
        },
    ),
    "idFatherImmutable": ManualSpec(
        status="New vs UK",
        source_kind="Java field",
        variable_name="idFatherImmutable",
        overrides={
            "Local": None,
            "Mod": "id",
            "Main attribute": "Father",
            "Other info": "Immutable",
            "Statistics type": None,
            "Description": "Immutable father ID used to preserve the original paternal link in simulation.",
            "Initial population": "No",
            "Module": "ID",
            "Notes": "No direct UK equivalent found in the reviewed workbook.",
        },
    ),
    "immutable_father_id": ManualSpec(
        status="New vs UK",
        source_kind="DB column",
        variable_name="idFatherImmutable",
        overrides={
            "Local": None,
            "Mod": "id",
            "Main attribute": "Father",
            "Other info": "Immutable",
            "Statistics type": None,
            "Description": "Database column backing the immutable father ID.",
            "Initial population": "No",
            "Module": "ID",
            "Notes": "Database column paired with idFatherImmutable.",
        },
    ),
    "staywparentsflag": ManualSpec(
        status="New vs UK",
        source_kind="Java field",
        variable_name="demStayParentFlag",
        review_required="Yes",
        review_note="Declared in Person.java but not referenced elsewhere in the current EU codebase; confirm whether to retain or remove in the refactor.",
        overrides={
            "Local": None,
            "Mod": "dem",
            "Main attribute": "StayParent",
            "Statistics type": "Flag",
            "Description": "Flag indicating that the person stays with parents.",
            "Initial population": "No",
            "Module": "Demography",
            "Notes": "New EU-only variable; no UK counterpart located in the reviewed workbook.",
        },
    ),
    "deh_c4": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="eduHighestC4",
        reference_old="deh_c3",
        review_required="Yes",
        review_note="EU uses a 4-state education enum including NotAssigned; confirm whether C4 should remain explicit in the refactor name.",
        overrides={
            "Other info": "C4",
            "Description": "Education - Highest Status\n 0 Not assigned\n 1 Low\n 2 Medium\n 3 High",
            "Notes": "Adapted from UK eduHighestC3 because SimPathsEU uses Education c4 / NotAssigned.",
        },
    ),
    "deh_c4_lag1": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="eduHighestC4L1",
        reference_old="deh_c3_lag1",
        review_required="Yes",
        review_note="Lag naming is clear, but the underlying education coding differs from the UK C3 variant.",
        overrides={
            "Other info": "C4",
            "Description": "Lag(1) of deh_c4",
            "Notes": "Adapted from UK eduHighestC3L1 because SimPathsEU uses Education c4 / NotAssigned.",
        },
    ),
    "dehm_c4": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="eduHighestMotherC4",
        reference_old="dehm_c3",
        review_required="Yes",
        review_note="EU uses a 4-state education enum including NotAssigned; confirm whether C4 should remain explicit in the refactor name.",
        overrides={
            "Other info": "C4",
            "Description": "Education - Mother's Highest Status\n 0 Not assigned\n 1 Low\n 2 Medium\n 3 High",
            "Notes": "Adapted from UK eduHighestMotherC3 because SimPathsEU uses Education c4 / NotAssigned.",
        },
    ),
    "dehf_c4": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="eduHighestFatherC4",
        reference_old="dehf_c3",
        review_required="Yes",
        review_note="EU uses a 4-state education enum including NotAssigned; confirm whether C4 should remain explicit in the refactor name.",
        overrides={
            "Other info": "C4",
            "Description": "Education - Father's Highest Status\n 0 Not assigned\n 1 Low\n 2 Medium\n 3 High",
            "Notes": "Adapted from UK eduHighestFatherC3 because SimPathsEU uses Education c4 / NotAssigned.",
        },
    ),
    "dehsp_c4_lag1": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="eduHighestPartnerC4L1",
        reference_old="dehsp_c3_lag1",
        review_required="Yes",
        review_note="EU uses a 4-state education enum including NotAssigned; confirm whether C4 should remain explicit in the refactor name.",
        overrides={
            "Other info": "C4",
            "Description": "Lag(1) of dehsp_c4",
            "Notes": "Adapted from UK eduHighestPartnerC3L1 because SimPathsEU uses Education c4 / NotAssigned.",
        },
    ),
    "deh_c4Local": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="i_eduHighestC4",
        reference_new="i_eduHighestC3",
        review_required="Yes",
        review_note="Local regression helper name follows the UK i_ prefix, but the EU education coding differs from UK C3.",
        overrides={
            "Local": "i_",
            "Other info": "C4",
            "Description": "Education - Highest Status local (c4 variant).",
            "Notes": "Adapted from UK i_eduHighestC3 because SimPathsEU uses Education c4 / NotAssigned.",
        },
    ),
    "ded_lag1": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="eduSpellFlagL1",
        reference_old="ded",
        overrides={
            "Periodicity": "L1",
            "Description": "Lag(1) of ded / in continuous education.",
            "Notes": "Derived from the UK eduSpellFlag naming pattern.",
        },
    ),
    "toRetire": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="demRtrdEnterFlag",
        reference_new="demRtrdEnterFlag",
        overrides={
            "Description": "Flag indicating that the individual is in the pool to retire this year.\n 0 No\n 1 Yes",
            "Java class": "Person.java",
            "Initial population": "No",
            "Notes": "Mapped to the UK retirement-entry flag naming pattern.",
        },
    ),
    "dhmGhq_lag1": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="healthPsyDstrssFlagL1",
        reference_new="healthPsyDstrssFlag",
        overrides={
            "Periodicity": "L1",
            "Description": "Lag(1) of dhm_ghq / psychological distress flag.",
            "Notes": "Derived from the UK healthPsyDstrssFlag naming pattern.",
        },
    ),
    "ydispPersInitial": ManualSpec(
        status="New vs UK",
        source_kind="Java field",
        variable_name="yDispPersMonthInitial",
        overrides={
            "Local": None,
            "Mod": "y",
            "Main attribute": None,
            "Level": "Pers",
            "Measurement": "Disp",
            "Other info": "Initial",
            "Statistics type": None,
            "Periodicity": "Month",
            "Description": "Personal monthly disposable income carried from the initial population.",
            "Initial population": "Yes",
            "Module": "Income",
            "Notes": "No direct UK equivalent found in the reviewed workbook.",
        },
    ),
    "ydisp_pers_initial": ManualSpec(
        status="New vs UK",
        source_kind="DB column",
        variable_name="yDispPersMonthInitial",
        overrides={
            "Local": None,
            "Mod": "y",
            "Level": "Pers",
            "Measurement": "Disp",
            "Other info": "Initial",
            "Periodicity": "Month",
            "Description": "Database column for personal monthly disposable income from the initial population.",
            "Initial population": "Yes",
            "Module": "Income",
            "Notes": "Database column paired with ydispPersInitial.",
        },
    ),
    "numberChildren02_lag1": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="demNChild0to2L1",
        reference_old="indicatorChildren03_lag1",
        review_required="Yes",
        review_note="UK uses an indicator-based lag for 0-2/3-under children here; EU stores a count, so the naming is adapted rather than copied.",
        overrides={
            "Local": None,
            "Mod": "dem",
            "Main attribute": "NChild",
            "Other info": "0to2",
            "Periodicity": "L1",
            "Description": "Lag(1) of the number of children aged 0-2 in the benefit unit.",
            "Initial population": "No",
            "Module": "Demography",
            "Notes": "Adapted from the UK child-count/child-indicator naming pattern.",
        },
    ),
    "numberChildrenAll_lag1": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="demNChildL1",
        reference_new="demNChild",
        review_required="Yes",
        review_note="No direct UK lag row exists for the benefit-unit child count; proposed from the UK demNChild base name.",
        overrides={
            "Local": None,
            "Mod": "dem",
            "Main attribute": "NChild",
            "Periodicity": "L1",
            "Description": "Lag(1) of the number of dependent children in the benefit unit.",
            "Initial population": "No",
            "Module": "Demography",
            "Notes": "Derived from the UK demNChild naming pattern.",
        },
    ),
    "dhhOwned_lag1": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="wealthPrptyFlagL1",
        reference_new="wealthPrptyFlag",
        review_required="Yes",
        review_note="UK has a person-level demPrptyFlagL1 row; for the EU benefit-unit variable, a wealth-module lag name is more consistent with wealthPrptyFlag.",
        overrides={
            "Local": None,
            "Mod": "wealth",
            "Main attribute": "Prpty",
            "Statistics type": "Flag",
            "Periodicity": "L1",
            "Description": "Lag(1) of dhhOwned / benefit-unit home-ownership flag.",
            "Initial population": "No",
            "Module": "Wealth",
            "Notes": "Derived from wealthPrptyFlag while keeping the lag suffix explicit.",
        },
    ),
    "edi_p50": ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="yHhDispEquivP50",
        reference_new="yHhDispEquivP50",
        overrides={
            "Java class": "Statistics.java",
            "Output file name": "File: Statistics",
            "Notes": "Mapped to the UK median equivalised disposable income naming pattern.",
        },
    ),
    "EDI_p50": ManualSpec(
        status="Adapted from UK",
        source_kind="DB column",
        variable_name="yHhDispEquivP50",
        reference_new="yHhDispEquivP50",
        overrides={
            "Java class": "Statistics.java",
            "Output file name": "File: Statistics",
            "Notes": "Database column paired with edi_p50.",
        },
    ),
    "sIndex_p50": ManualSpec(
        status="New vs UK",
        source_kind="Java field",
        variable_name="statSIndexP50",
        reference_new="statSIndex",
        overrides={
            "Local": None,
            "Mod": "stat",
            "Main attribute": "SIndex",
            "Statistics type": "P50",
            "Description": "Median S index.",
            "Java class": "Statistics.java",
            "Initial population": "No",
            "Output file name": "File: Statistics",
            "Module": "Statistical display",
            "Notes": "UK workbook contains statSIndex/statSIndexNormal but not an S-index median row.",
        },
    ),
    "SIndex_p50": ManualSpec(
        status="New vs UK",
        source_kind="DB column",
        variable_name="statSIndexP50",
        reference_new="statSIndex",
        overrides={
            "Local": None,
            "Mod": "stat",
            "Main attribute": "SIndex",
            "Statistics type": "P50",
            "Description": "Database column for the median S index.",
            "Java class": "Statistics.java",
            "Initial population": "No",
            "Output file name": "File: Statistics",
            "Module": "Statistical display",
            "Notes": "Database column paired with sIndex_p50.",
        },
    ),
}

CLASS_MANUAL_SPECS = {
    ("Person.java", "socialCareProvision"): ManualSpec(
        status="New vs UK",
        source_kind="Java field",
        variable_name="careProvidedToC4",
        match_basis="Manual review of EU-only care-provision categories",
        review_required="Yes",
        review_note="Person.socialCareProvision is a 4-category enum in SimPathsEU, not a simple provided-care flag as in the reviewed UK workbook.",
        overrides={
            "Local": None,
            "Mod": "care",
            "Main attribute": None,
            "Level": None,
            "Measurement": None,
            "Other info": "ProvidedTo",
            "Statistics type": "C4",
            "Periodicity": None,
            "Description": "Social-care provision category\n 0 None\n 1 Only partner\n 2 Partner and other\n 3 Only other",
            "Initial population": "No",
            "Module": "Social care",
            "Notes": "No exact UK analogue found; the UK workbook row for socialCareProvision is a benefit-unit flag rather than a person-level category.",
        },
    ),
    ("Person.java", "socialCareProvision_lag1"): ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="careProvidedToC4L1",
        reference_old="socialCareProvision_lag1",
        match_basis="Manual adaptation of care-provision lag naming",
        review_required="Yes",
        review_note="Lag naming follows the EU person-level category variable rather than the UK flag-style naming.",
        overrides={
            "Other info": "ProvidedTo",
            "Statistics type": "C4",
            "Periodicity": "L1",
            "Description": "Lag(1) of social-care provision category.",
            "Notes": "Adapted from the reviewed UK lag row because SimPathsEU stores a category, not a flag.",
        },
    ),
    ("Person.java", "idFather"): ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="idFather",
        reference_new="idFather",
        match_basis="Manual correction of UK workbook inconsistency",
        overrides={
            "Main attribute": "Father",
            "Description": "Father ID",
            "Module": "ID",
            "Notes": "Corrected against a conflicting UK workbook row where old name idFather was assigned the idMother variable name.",
        },
    ),
    ("Statistics.java", "ydses_p60"): ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="yHhQuintilesC5P60",
        reference_old="ydses_p40",
        match_basis="Manual correction of UK workbook inconsistency",
        overrides={
            "Statistics type": "C5P60",
            "Variable name (concatenate)": "yHhQuintilesC5P60",
            "Variable name": "yHhQuintilesC5P60",
            "Description": "Percentile (60) of ydses_c5",
            "Java class": "Statistics.java",
            "Output file name": "File: Statistics",
            "Notes": "Corrected because the reviewed UK workbook labels ydses_p60 as yHhQuintilesC5P80.",
        },
    ),
    ("Statistics.java", "edi_p50"): ManualSpec(
        status="Adapted from UK",
        source_kind="Java field",
        variable_name="yHhDispEquivP50Calc",
        reference_new="yHhDispEquivP50",
        match_basis="Manual review of duplicate EDI median fields",
        review_required="Yes",
        review_note="Statistics.java stores both medianEquivalisedHouseholdDisposableIncome and edi_p50; confirm whether both should survive the refactor or be consolidated.",
        overrides={
            "Other info": "Calc",
            "Variable name (concatenate)": "yHhDispEquivP50Calc",
            "Variable name": "yHhDispEquivP50Calc",
            "Description": "Percentile-calculated P50 of equivalised household disposable income.",
            "Java class": "Statistics.java",
            "Output file name": "File: Statistics",
            "Notes": "Separated from the main yHhDispEquivP50 row because SimPathsEU stores two median-EDI fields in Statistics.java.",
        },
    ),
}


ALIGNMENT_SPECS = {
    "retirementAdjustmentFactor": ("labRtrdAdj", "lab", "Rtrd", "Adj", None, "Retirement adjustment factor."),
    "retirement_adj_factor": ("labRtrdAdj", "lab", "Rtrd", "Adj", None, "Database column for the retirement adjustment factor."),
    "disabilityAdjustmentFactor": ("healthDsblAdj", "health", "Dsbl", "Adj", None, "Disability adjustment factor."),
    "disability_adj_factor": ("healthDsblAdj", "health", "Dsbl", "Adj", None, "Database column for the disability adjustment factor."),
    "retirementShareSimulated": ("labRtrdSimShare", "lab", "Rtrd", None, "Sim", "Simulated retirement share."),
    "retirement_share_sim": ("labRtrdSimShare", "lab", "Rtrd", None, "Sim", "Database column for the simulated retirement share."),
    "retirementShareTarget": ("labRtrdTargetShare", "lab", "Rtrd", None, "Target", "Target retirement share."),
    "retirement_share_tgt": ("labRtrdTargetShare", "lab", "Rtrd", None, "Target", "Database column for the target retirement share."),
    "disabilityShareSimulated": ("healthDsblSimShare", "health", "Dsbl", None, "Sim", "Simulated disability share."),
    "disability_share_sim": ("healthDsblSimShare", "health", "Dsbl", None, "Sim", "Database column for the simulated disability share."),
    "disabilityShareTarget": ("healthDsblTargetShare", "health", "Dsbl", None, "Target", "Target disability share."),
    "disability_share_tgt": ("healthDsblTargetShare", "health", "Dsbl", None, "Target", "Database column for the target disability share."),
    "inSchoolAdjustmentFactor": ("eduSpellAdj", "edu", "Spell", "Adj", None, "In-school adjustment factor."),
    "in_school_adj_factor": ("eduSpellAdj", "edu", "Spell", "Adj", None, "Database column for the in-school adjustment factor."),
    "inSchoolShareSimulated": ("eduSpellSimShare", "edu", "Spell", None, "Sim", "Simulated in-school share."),
    "in_school_share_sim": ("eduSpellSimShare", "edu", "Spell", None, "Sim", "Database column for the simulated in-school share."),
    "inSchoolShareTarget": ("eduSpellTargetShare", "edu", "Spell", None, "Target", "Target in-school share."),
    "in_school_share_tgt": ("eduSpellTargetShare", "edu", "Spell", None, "Target", "Database column for the target in-school share."),
    "utilityAdjustmentFactorACMale": ("demUtilAdjAcMale", "dem", "Util", "Adj", "AcMale", "Utility adjustment factor adult-child males."),
    "utility_adj_factor_ac_male": ("demUtilAdjAcMale", "dem", "Util", "Adj", "AcMale", "Database column for the adult-child male utility adjustment factor."),
    "utilityAdjustmentFactorACFemale": ("demUtilAdjAcFemale", "dem", "Util", "Adj", "AcFemale", "Utility adjustment factor adult-child females."),
    "utility_adj_factor_ac_female": ("demUtilAdjAcFemale", "dem", "Util", "Adj", "AcFemale", "Database column for the adult-child female utility adjustment factor."),
    "utilityAdjustmentFactorMaleWithDep": ("demUtilAdjMaleWithDep", "dem", "Util", "Adj", "MaleWithDep", "Utility adjustment factor males with dependants."),
    "utility_adj_factor_male_with_dep": ("demUtilAdjMaleWithDep", "dem", "Util", "Adj", "MaleWithDep", "Database column for the male-with-dependants utility adjustment factor."),
    "utilityAdjustmentFactorFemaleWithDep": ("demUtilAdjFemaleWithDep", "dem", "Util", "Adj", "FemaleWithDep", "Utility adjustment factor females with dependants."),
    "utility_adj_factor_female_with_dep": ("demUtilAdjFemaleWithDep", "dem", "Util", "Adj", "FemaleWithDep", "Database column for the female-with-dependants utility adjustment factor."),
    "employedShareSimSingleMales": ("labEmpSimShareSingleM", "lab", "Emp", None, "SingleM", "Simulated employed share for single males."),
    "employed_share_sim_smales": ("labEmpSimShareSingleM", "lab", "Emp", None, "SingleM", "Database column for the simulated employed share of single males."),
    "employedShareTgtSingleMales": ("labEmpTargetShareSingleM", "lab", "Emp", None, "SingleMTarget", "Target employed share for single males."),
    "employed_share_tgt_smales": ("labEmpTargetShareSingleM", "lab", "Emp", None, "SingleMTarget", "Database column for the target employed share of single males."),
    "employedShareSimSingleFemales": ("labEmpSimShareSingleF", "lab", "Emp", None, "SingleF", "Simulated employed share for single females."),
    "employed_share_sim_sfemales": ("labEmpSimShareSingleF", "lab", "Emp", None, "SingleF", "Database column for the simulated employed share of single females."),
    "employedShareTgtSingleFemales": ("labEmpTargetShareSingleF", "lab", "Emp", None, "SingleFTarget", "Target employed share for single females."),
    "employed_share_tgt_sfemales": ("labEmpTargetShareSingleF", "lab", "Emp", None, "SingleFTarget", "Database column for the target employed share of single females."),
    "employedShareSimCouples": ("labEmpSimShareCouple", "lab", "Emp", None, "Couple", "Simulated employed share for couples."),
    "employed_share_sim_couples": ("labEmpSimShareCouple", "lab", "Emp", None, "Couple", "Database column for the simulated employed share of couples."),
    "employedShareTgtCouples": ("labEmpTargetShareCouple", "lab", "Emp", None, "CoupleTarget", "Target employed share for couples."),
    "employed_share_tgt_couples": ("labEmpTargetShareCouple", "lab", "Emp", None, "CoupleTarget", "Database column for the target employed share of couples."),
    "employedShareSimACMale": ("labEmpSimShareAcMale", "lab", "Emp", None, "AcMale", "Simulated employed share for adult-child males."),
    "employed_share_sim_ac_male": ("labEmpSimShareAcMale", "lab", "Emp", None, "AcMale", "Database column for the simulated employed share of adult-child males."),
    "employedShareTgtACMale": ("labEmpTargetShareAcMale", "lab", "Emp", None, "AcMaleTarget", "Target employed share for adult-child males."),
    "employed_share_tgt_ac_male": ("labEmpTargetShareAcMale", "lab", "Emp", None, "AcMaleTarget", "Database column for the target employed share of adult-child males."),
    "employedShareSimACFemale": ("labEmpSimShareAcFemale", "lab", "Emp", None, "AcFemale", "Simulated employed share for adult-child females."),
    "employed_share_sim_ac_female": ("labEmpSimShareAcFemale", "lab", "Emp", None, "AcFemale", "Database column for the simulated employed share of adult-child females."),
    "employedShareTgtACFemale": ("labEmpTargetShareAcFemale", "lab", "Emp", None, "AcFemaleTarget", "Target employed share for adult-child females."),
    "employed_share_tgt_ac_female": ("labEmpTargetShareAcFemale", "lab", "Emp", None, "AcFemaleTarget", "Database column for the target employed share of adult-child females."),
    "employedShareSimMaleWithDep": ("labEmpSimShareMaleWithDep", "lab", "Emp", None, "MaleWithDep", "Simulated employed share for males with dependants."),
    "employed_share_sim_male_with_dep": ("labEmpSimShareMaleWithDep", "lab", "Emp", None, "MaleWithDep", "Database column for the simulated employed share of males with dependants."),
    "employedShareTgtMaleWithDep": ("labEmpTargetShareMaleWithDep", "lab", "Emp", None, "MaleWithDepTarget", "Target employed share for males with dependants."),
    "employed_share_tgt_male_with_dep": ("labEmpTargetShareMaleWithDep", "lab", "Emp", None, "MaleWithDepTarget", "Database column for the target employed share of males with dependants."),
    "employedShareSimFemaleWithDep": ("labEmpSimShareFemaleWithDep", "lab", "Emp", None, "FemaleWithDep", "Simulated employed share for females with dependants."),
    "employed_share_sim_female_with_dep": ("labEmpSimShareFemaleWithDep", "lab", "Emp", None, "FemaleWithDep", "Database column for the simulated employed share of females with dependants."),
    "employedShareTgtFemaleWithDep": ("labEmpTargetShareFemaleWithDep", "lab", "Emp", None, "FemaleWithDepTarget", "Target employed share for females with dependants."),
    "employed_share_tgt_female_with_dep": ("labEmpTargetShareFemaleWithDep", "lab", "Emp", None, "FemaleWithDepTarget", "Database column for the target employed share of females with dependants."),
}


CONTROL_MANUAL = {
    "TimeSeriesVariable": {
        "CareProvisionAdjustment": ("careAdj", "Adapted from UK", "Social care adjustment factor (time-series control).", "No", ""),
        "CarerWageRate": ("careWageRate", "New vs UK", "Carer wage rate used for formal-care cost calculations.", "No", ""),
        "FixedRetirementAge": ("labRtrdAgeFixed", "New vs UK", "Fixed retirement age schedule.", "No", ""),
        "GDP": ("statGdp", "New vs UK", "GDP time-series control.", "No", ""),
        "HighEducationRate": ("eduHighRate", "New vs UK", "Target rate for high education attainment.", "No", ""),
        "Inflation": ("statInflation", "New vs UK", "Inflation time-series index.", "No", ""),
        "LowEducationRate": ("eduLowRate", "New vs UK", "Target rate for low education attainment.", "No", ""),
        "PartnershipAdjustment": ("demPartnerAdj", "Adapted from UK", "Partnership adjustment factor control.", "No", ""),
        "FertilityAdjustment": ("demFertAdj", "Adapted from UK", "Fertility adjustment factor control.", "No", ""),
        "DisabilityAdjustment": ("healthDsblAdj", "New vs UK", "Disability adjustment factor control.", "No", ""),
        "UtilityAdjustment": ("demUtilAdj", "New vs UK", "Top-level utility adjustment factor.", "Yes", "No UK row exists for the aggregate utility-adjustment control; confirm whether the aggregate control should remain separate from subgroup controls."),
        "UtilityAdjustmentSingleMales": ("demUtilAdjSingleM", "Adapted from UK", "Utility adjustment factor for single males.", "No", ""),
        "UtilityAdjustmentACMales": ("demUtilAdjAcMale", "New vs UK", "Utility adjustment factor for adult-child males.", "Yes", "Adult-child subgroup naming is inferred because the UK workbook has no direct AC row."),
        "UtilityAdjustmentSingleFemales": ("demUtilAdjSingleF", "Adapted from UK", "Utility adjustment factor for single females.", "No", ""),
        "UtilityAdjustmentACFemales": ("demUtilAdjAcFemale", "New vs UK", "Utility adjustment factor for adult-child females.", "Yes", "Adult-child subgroup naming is inferred because the UK workbook has no direct AC row."),
        "UtilityAdjustmentCouples": ("demUtilAdjCouple", "Adapted from UK", "Utility adjustment factor for couples.", "No", ""),
        "UtilityAdjustmentMaleWithDep": ("demUtilAdjMaleWithDep", "New vs UK", "Utility adjustment factor for males with dependants.", "Yes", "With-dependants subgroup naming is inferred because the UK workbook has no direct row."),
        "UtilityAdjustmentFemaleWithDep": ("demUtilAdjFemaleWithDep", "New vs UK", "Utility adjustment factor for females with dependants.", "Yes", "With-dependants subgroup naming is inferred because the UK workbook has no direct row."),
        "RetirementAdjustment": ("labRtrdAdj", "New vs UK", "Retirement adjustment factor control.", "No", ""),
        "InSchoolAdjustment": ("eduSpellAdj", "New vs UK", "In-school adjustment factor control.", "No", ""),
        "WageGrowth": ("labWageGrowth", "New vs UK", "Wage-growth time-series index.", "No", ""),
    },
    "AlignmentVariable": {
        "PartnershipAlignment": ("demPartnerAlign", "New vs UK", "Partnership alignment control enum.", "No", ""),
        "FertilityAlignment": ("demFertAlign", "New vs UK", "Fertility alignment control enum.", "No", ""),
        "RetirementAlignment": ("labRtrdAlign", "New vs UK", "Retirement alignment control enum.", "No", ""),
        "DisabilityAlignment": ("healthDsblAlign", "New vs UK", "Disability alignment control enum.", "No", ""),
    },
}


def read_variables_sheet(workbook):
    ws = workbook["Variables"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    row_dicts = []
    for row_number, row in enumerate(rows[1:], start=2):
        values = list(row) + [None] * (len(header) - len(row))
        row_dict = {header[index]: values[index] for index in range(len(header))}
        row_dict["_row_number"] = row_number
        row_dicts.append(row_dict)
    return header, row_dicts


def build_row_maps(uk_rows):
    by_old_exact = {}
    by_old_lower = {}
    by_new = {}
    for row in uk_rows:
        old_name = row.get("SimPaths old name")
        new_name = row.get("Variable name")
        if old_name:
            old_text = str(old_name).strip()
            by_old_exact.setdefault(old_text, []).append(row)
            by_old_lower.setdefault(old_text.lower(), []).append(row)
        if new_name:
            by_new.setdefault(str(new_name).strip().lower(), []).append(row)
    return by_old_exact, by_old_lower, by_new


def strip_comments(text: str) -> str:
    return COMMENT_RE.sub("", text)


def parse_fields(java_path: Path) -> list[Candidate]:
    text = strip_comments(java_path.read_text())
    candidates = []
    order = 0
    for match in FIELD_RE.finditer(text):
        order += 1
        name = match.group("name")
        column_match = COLUMN_RE.search(match.group(0))
        column_name = column_match.group(1) if column_match else None
        candidates.append(
            Candidate(
                source_name=name,
                source_kind="Java field",
                java_class=java_path.name,
                java_path=java_path,
                order=order,
                sibling_name=column_name,
            )
        )
    return candidates


def parse_enum_constants(java_path: Path, enum_name: str) -> list[str]:
    text = strip_comments(java_path.read_text())
    match = re.search(r"\benum\s+" + re.escape(enum_name) + r"\s*\{", text)
    if not match:
        return []
    start = match.end()
    depth = 1
    cursor = start
    while cursor < len(text) and depth > 0:
        if text[cursor] == "{":
            depth += 1
        elif text[cursor] == "}":
            depth -= 1
        cursor += 1
    body = text[start : cursor - 1]
    constants_section = body.split(";", 1)[0]
    constants = []
    for raw_part in constants_section.split(","):
        token = raw_part.strip()
        if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", token):
            constants.append(token)
    return constants


def candidate_is_excluded(candidate: Candidate) -> str | None:
    if candidate.source_name in EXCLUDED_REASONS:
        return EXCLUDED_REASONS[candidate.source_name]
    lower_name = candidate.source_name.lower()
    if candidate.java_class == "Processed.java":
        return "Metadata record field"
    if lower_name.endswith("series"):
        return "Series wrapper helper"
    if lower_name.endswith("cachekey") or lower_name.endswith("cacheyear"):
        return "Cache field"
    if lower_name.startswith("cached"):
        return "Cache field"
    if lower_name.endswith("map") and candidate.source_kind == "Java field":
        return "Helper map/cache"
    if candidate.source_name in {"benefitUnit", "household", "members"}:
        return "Entity association"
    return None


def make_base_row(header: list[str]) -> dict[str, object]:
    row = {column: None for column in header}
    row["Variable name (concatenate)"] = None
    row["Variable name"] = None
    row["Description"] = None
    row["Java class"] = None
    row["Initial population"] = "No"
    row["Module"] = None
    row["Output file name"] = None
    row["Notes"] = None
    row["SimPaths old name"] = None
    return row


def clone_row(template_row: dict[str, object], header: list[str]) -> dict[str, object]:
    row = {column: template_row.get(column) for column in header}
    return row


def choose_template_by_new(
    new_name: str,
    class_hint: str,
    uk_by_new: dict[str, list[dict[str, object]]],
) -> dict[str, object] | None:
    options = uk_by_new.get(new_name.lower(), [])
    if not options:
        return None
    for option in options:
        if option.get("Java class") == class_hint:
            return option
    return options[0]


def choose_template_by_old(
    old_name: str,
    class_hint: str,
    uk_by_old_exact: dict[str, list[dict[str, object]]],
    uk_by_old_lower: dict[str, list[dict[str, object]]],
) -> tuple[dict[str, object] | None, str | None]:
    if not old_name:
        return None, None

    best_template = None
    best_basis = None
    best_score = -1

    search_spaces = [
        ("UK old name exact", uk_by_old_exact.get(old_name, []), 10),
        ("UK old name lower-case", uk_by_old_lower.get(old_name.lower(), []), 5),
    ]

    for basis, options, base_score in search_spaces:
        for option in options:
            score = base_score
            if option.get("Java class") == class_hint:
                score += 3
            if str(option.get("SimPaths old name") or "") == old_name:
                score += 1
            if score > best_score:
                best_template = option
                best_basis = basis
                best_score = score

    return best_template, best_basis


def find_template_for_candidate(
    candidate: Candidate,
    uk_by_old_exact: dict[str, list[dict[str, object]]],
    uk_by_old_lower: dict[str, list[dict[str, object]]],
) -> tuple[dict[str, object] | None, str | None]:
    search_items = [
        ("EU field", candidate.source_name),
        ("EU DB alias", candidate.sibling_name),
    ]

    best_template = None
    best_basis = None
    best_score = -1

    for label, raw_name in search_items:
        if not raw_name:
            continue
        template, basis = choose_template_by_old(raw_name, candidate.java_class, uk_by_old_exact, uk_by_old_lower)
        if not template:
            continue
        score = 0
        if label == "EU field":
            score += 4
        if basis == "UK old name exact":
            score += 2
        if template.get("Java class") == candidate.java_class:
            score += 3
        if score > best_score:
            best_template = template
            best_basis = f"{label} via {basis}"
            best_score = score

    return best_template, best_basis


def build_manual_row(
    header: list[str],
    uk_by_old_exact: dict[str, list[dict[str, object]]],
    uk_by_old_lower: dict[str, list[dict[str, object]]],
    uk_by_new: dict[str, list[dict[str, object]]],
    candidate: Candidate,
    spec: ManualSpec,
) -> tuple[dict[str, object], int | None]:
    if spec.reference_old:
        template, _ = choose_template_by_old(spec.reference_old, candidate.java_class, uk_by_old_exact, uk_by_old_lower)
    elif spec.reference_new:
        template = choose_template_by_new(spec.reference_new, candidate.java_class, uk_by_new)
    else:
        template = None

    if template:
        row = clone_row(template, header)
        template_row_number = template["_row_number"]
    else:
        row = make_base_row(header)
        template_row_number = None

    if spec.variable_name:
        row["Variable name (concatenate)"] = spec.variable_name
        row["Variable name"] = spec.variable_name
    row["Java class"] = candidate.java_class
    row["SimPaths old name"] = candidate.source_name
    row["EU DB column / alias"] = candidate.sibling_name if candidate.sibling_name != candidate.source_name else None
    if candidate.java_class in {"Statistics.java", "Statistics2.java", "AlignmentAdjustmentFactors.java"}:
        row["Output file name"] = f"File: {candidate.java_class.replace('.java', '')}"

    if spec.overrides:
        row.update(spec.overrides)

    row["Comparison status"] = spec.status
    row["Match basis"] = spec.match_basis
    row["UK reference"] = spec.reference_new or spec.reference_old
    row["Needs review"] = spec.review_required
    row["Review note"] = spec.review_note
    return row, template_row_number


def build_alignment_row(header: list[str], candidate: Candidate) -> dict[str, object]:
    variable_name, mod, main_attribute, measurement, other_info, description = ALIGNMENT_SPECS[candidate.source_name]
    row = make_base_row(header)
    row["Local"] = None
    row["Mod"] = mod
    row["Main attribute"] = main_attribute
    row["Measurement"] = measurement
    row["Other info"] = other_info
    row["Variable name (concatenate)"] = variable_name
    row["Variable name"] = variable_name
    row["Description"] = description
    row["Java class"] = candidate.java_class
    row["Initial population"] = "No"
    row["Output file name"] = "File: AlignmentAdjustmentFactors"
    row["Module"] = {
        "dem": "Demography",
        "edu": "Education",
        "health": "Health",
        "lab": "Labour",
    }[mod]
    row["Notes"] = "New SimPathsEU alignment output relative to the reviewed UK workbook."
    row["SimPaths old name"] = candidate.source_name
    row["EU DB column / alias"] = candidate.sibling_name if candidate.sibling_name != candidate.source_name else None
    row["Comparison status"] = "New vs UK"
    row["Match basis"] = "Manual review of SimPathsEU alignment outputs"
    row["UK reference"] = None
    row["Needs review"] = "Yes" if "Ac" in variable_name or "WithDep" in variable_name else "No"
    row["Review note"] = (
        "Subgroup naming is inferred because the reviewed UK workbook does not contain this alignment subgroup."
        if row["Needs review"] == "Yes"
        else ""
    )
    return row


def build_exact_row(
    header: list[str],
    template: dict[str, object],
    candidate: Candidate,
    status: str,
    match_basis: str,
) -> dict[str, object]:
    row = clone_row(template, header)
    row["Java class"] = candidate.java_class
    row["SimPaths old name"] = candidate.source_name
    row["EU DB column / alias"] = candidate.sibling_name if candidate.sibling_name != candidate.source_name else None
    row["Comparison status"] = status
    row["Match basis"] = match_basis
    row["UK reference"] = template.get("Variable name")
    row["Needs review"] = "No"
    row["Review note"] = ""
    if candidate.java_class in {"Statistics.java", "Statistics2.java", "AlignmentAdjustmentFactors.java"}:
        row["Output file name"] = f"File: {candidate.java_class.replace('.java', '')}"
    return row


def lookup_manual_spec(candidate: Candidate) -> ManualSpec | None:
    return CLASS_MANUAL_SPECS.get((candidate.java_class, candidate.source_name)) or MANUAL_SPECS.get(candidate.source_name)


def row_fill_for_status(status: str):
    if status == "Matched UK":
        return MATCHED_FILL
    if status == "Adapted from UK":
        return ADAPTED_FILL
    if status == "New vs UK":
        return NEW_FILL
    return None


def autofit_columns(ws):
    for column_cells in ws.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = max(len(value) for value in values)
        width = min(max(max_length + 2, 10), 40)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def reset_sheet(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def write_header(ws, header: list[str]):
    for index, column_name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=index, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = HEADER_STYLE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_rows(ws, header: list[str], rows: Iterable[dict[str, object]]):
    for row_number, row in enumerate(rows, start=2):
        for column_number, column_name in enumerate(header, start=1):
            ws.cell(row=row_number, column=column_number, value=row.get(column_name))
        fill = row_fill_for_status(str(row.get("Comparison status")))
        if fill:
            for column_number in range(1, len(header) + 1):
                ws.cell(row=row_number, column=column_number).fill = fill
        if row.get("Needs review") == "Yes":
            for column_name in ("Needs review", "Review note"):
                column_index = header.index(column_name) + 1
                ws.cell(row=row_number, column=column_index).fill = REVIEW_FILL
        for column_name in ("Description", "Notes", "Review note"):
            column_index = header.index(column_name) + 1
            ws.cell(row=row_number, column=column_index).alignment = Alignment(wrap_text=True, vertical="top")
    ws.auto_filter.ref = ws.dimensions


def sort_codebook_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    def key(row):
        mod = str(row.get("Mod") or "")
        variable_name = str(row.get("Variable name") or "")
        java_class = str(row.get("Java class") or "")
        old_name = str(row.get("SimPaths old name") or "")
        return (MODULE_ORDER.get(mod, 99), mod, variable_name.lower(), java_class.lower(), old_name.lower())

    return sorted(rows, key=key)


def update_cover_sheet(workbook):
    ws = workbook["Cover"]
    ws["B2"] = "SimPathsEU Codebook"
    ws["C4"] = date.today()
    ws["C5"] = "Codex review based on SimPaths_Variable_CodebookUK"
    ws["B7"] = "Workbook contents"
    ws["B8"] = "Variables: one logical SimPathsEU field per row, with DB aliases attached to the same row"
    ws["B9"] = "EU Review Summary: scope, counts, highlight legend, and review totals"
    ws["B10"] = "Collision Review: proposed names still shared across multiple EU fields/classes"
    ws["B11"] = "UK Review Issues: inconsistencies detected in the reviewed UK workbook"
    ws["B12"] = "Controls: time-series and alignment enums reviewed for refactoring"
    ws["B13"] = "UK Only: variables present in the reviewed UK workbook but not matched in SimPathsEU"
    ws["B14"] = "Excluded Internals: infrastructure/cache/metadata fields excluded from the codebook"
    ws["B15"] = "Deferred Inventory: regression and validation enums reviewed but left outside the main UK-style codebook pass"


def build_summary_rows(
    included_rows: list[dict[str, object]],
    uk_only_rows: list[dict[str, object]],
    excluded_rows: list[dict[str, object]],
    deferred_rows: list[dict[str, object]],
    control_rows: list[dict[str, object]],
    collision_rows: list[dict[str, object]],
    uk_issue_rows: list[dict[str, object]],
):
    counts = {
        "Core variables included": len(included_rows),
        "Matched UK": sum(1 for row in included_rows if row["Comparison status"] == "Matched UK"),
        "Adapted from UK": sum(1 for row in included_rows if row["Comparison status"] == "Adapted from UK"),
        "New vs UK": sum(1 for row in included_rows if row["Comparison status"] == "New vs UK"),
        "Needs review": sum(1 for row in included_rows if row["Needs review"] == "Yes"),
        "Remaining EU name-collision groups": len({row["Variable name"] for row in collision_rows}),
        "UK workbook review issues logged": len(uk_issue_rows),
        "Controls reviewed": len(control_rows),
        "UK-only rows": len(uk_only_rows),
        "Excluded internal/metadata fields": len(excluded_rows),
        "Deferred regression/validation enum constants": len(deferred_rows),
    }

    lines = [
        ("Generated on", date.today().isoformat()),
        ("Scope", "Person, BenefitUnit, Household, Statistics, Statistics2, AlignmentAdjustmentFactors; UK workbook reviewed as naming reference."),
        ("Coverage note", "Core variables are codebooked one logical field per row. Time-series/alignment enums are reviewed in Controls. Regression/validation enums are inventoried separately in Deferred Inventory."),
    ]
    for key, value in counts.items():
        lines.append((key, value))
    lines.extend(
        [
            ("Highlight legend", None),
            ("Green rows", "Direct UK mapping reused for SimPathsEU."),
            ("Blue rows", "UK naming pattern reused but adapted for an EU variant."),
            ("Yellow rows", "New SimPathsEU variable relative to the reviewed UK workbook."),
            ("Orange review cells", "Name needs confirmation because the UK workbook had no exact analogue or the EU semantics differ."),
        ]
    )
    return lines


def build_collision_rows(included_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    grouped = {}
    for row in included_rows:
        grouped.setdefault(row["Variable name"], []).append(row)

    collisions = []
    for variable_name, rows in grouped.items():
        if len(rows) < 2:
            continue
        for row in rows:
            collisions.append(
                {
                    "Variable name": variable_name,
                    "Java class": row.get("Java class"),
                    "SimPaths old name": row.get("SimPaths old name"),
                    "EU DB column / alias": row.get("EU DB column / alias"),
                    "Comparison status": row.get("Comparison status"),
                    "Needs review": row.get("Needs review"),
                    "Review note": row.get("Review note") or "Shared proposed name across multiple EU fields/classes; review if global uniqueness is required for the refactor.",
                }
            )
    return sorted(collisions, key=lambda item: (str(item["Variable name"]), str(item["Java class"]), str(item["SimPaths old name"])))


def build_uk_review_issues(uk_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    issues = []

    for row in uk_rows:
        old_name = str(row.get("SimPaths old name") or "")
        variable_name = str(row.get("Variable name") or "")
        if old_name == "idFather" and variable_name == "idMother":
            issues.append(
                {
                    "Issue type": "Incorrect mapping",
                    "Variable name": variable_name,
                    "SimPaths old name": old_name,
                    "Java class": row.get("Java class"),
                    "Description": row.get("Description"),
                    "Review note": "UK workbook row uses idMother as the proposed name for old name idFather despite a Father ID description.",
                }
            )
        if old_name == "ydses_p60" and variable_name == "yHhQuintilesC5P80":
            issues.append(
                {
                    "Issue type": "Incorrect statistic label",
                    "Variable name": variable_name,
                    "SimPaths old name": old_name,
                    "Java class": row.get("Java class"),
                    "Description": row.get("Description"),
                    "Review note": "UK workbook labels the 60th percentile row as C5P80.",
                }
            )

    duplicate_counts = {}
    for row in uk_rows:
        variable_name = row.get("Variable name")
        if variable_name:
            duplicate_counts.setdefault(variable_name, []).append(row)

    for variable_name, rows in duplicate_counts.items():
        if len(rows) < 2:
            continue
        issues.append(
            {
                "Issue type": "Non-unique variable name",
                "Variable name": variable_name,
                "SimPaths old name": "; ".join(str(row.get("SimPaths old name")) for row in rows),
                "Java class": "; ".join(str(row.get("Java class")) for row in rows),
                "Description": rows[0].get("Description"),
                "Review note": f"Reviewed UK workbook uses this proposed name for {len(rows)} different rows/classes.",
            }
        )

    return sorted(issues, key=lambda item: (str(item["Issue type"]), str(item["Variable name"])))


def create_or_replace_sheet(workbook, title: str):
    if title in workbook.sheetnames:
        index = workbook.sheetnames.index(title)
        del workbook[title]
        return workbook.create_sheet(title, index)
    return workbook.create_sheet(title)


def clone_sheet_layout(source_ws, target_ws):
    target_ws.sheet_format.defaultColWidth = source_ws.sheet_format.defaultColWidth
    target_ws.sheet_format.defaultRowHeight = source_ws.sheet_format.defaultRowHeight
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.auto_filter.ref = source_ws.auto_filter.ref
    target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
    target_ws.sheet_properties.tabColor = source_ws.sheet_properties.tabColor

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    for key, dimension in source_ws.column_dimensions.items():
        target_dimension = target_ws.column_dimensions[key]
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden
        target_dimension.bestFit = dimension.bestFit

    for key, dimension in source_ws.row_dimensions.items():
        target_dimension = target_ws.row_dimensions[key]
        target_dimension.height = dimension.height
        target_dimension.hidden = dimension.hidden
        target_dimension.outlineLevel = dimension.outlineLevel


def clone_sheet_cells(source_ws, target_ws):
    for row in source_ws.iter_rows():
        for cell in row:
            target = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                target.font = copy.copy(cell.font)
                target.fill = copy.copy(cell.fill)
                target.border = copy.copy(cell.border)
                target.alignment = copy.copy(cell.alignment)
                target.number_format = cell.number_format
                target.protection = copy.copy(cell.protection)
            if cell.hyperlink:
                target._hyperlink = copy.copy(cell.hyperlink)
            if cell.comment:
                target.comment = copy.copy(cell.comment)


def copy_reference_sheet(source_workbook, target_workbook, title: str):
    source_ws = source_workbook[title]
    target_ws = target_workbook.create_sheet(title)
    clone_sheet_cells(source_ws, target_ws)
    clone_sheet_layout(source_ws, target_ws)
    return target_ws


def main():
    uk_workbook = load_workbook(UK_CODEBOOK)
    base_header, uk_rows = read_variables_sheet(uk_workbook)
    uk_by_old_exact, uk_by_old_lower, uk_by_new = build_row_maps(uk_rows)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    copy_reference_sheet(uk_workbook, workbook, "Cover")
    variables_ws = workbook.create_sheet("Variables")
    for title in ("Rules", "Modules", "Abbreviations", "Coding Style", "Country Specific Vars"):
        copy_reference_sheet(uk_workbook, workbook, title)

    extended_header = base_header + [
        "EU DB column / alias",
        "Comparison status",
        "Match basis",
        "UK reference",
        "Needs review",
        "Review note",
    ]

    included_rows = []
    excluded_rows = []
    used_uk_rows = set()

    all_candidates = []
    for java_file in CORE_JAVA_FILES:
        all_candidates.extend(parse_fields(java_file))

    for candidate in all_candidates:
        exclusion_reason = candidate_is_excluded(candidate)
        if exclusion_reason:
            excluded_rows.append(
                {
                    "Java class": candidate.java_class,
                    "Raw name": candidate.source_name,
                    "Source kind": "Java field",
                    "Reason": exclusion_reason,
                }
            )
            continue

        manual_spec = lookup_manual_spec(candidate)
        if manual_spec:
            row, template_row_number = build_manual_row(
                extended_header,
                uk_by_old_exact,
                uk_by_old_lower,
                uk_by_new,
                candidate,
                manual_spec,
            )
            included_rows.append(row)
            if template_row_number:
                used_uk_rows.add(template_row_number)
            continue

        if candidate.source_name in ALIGNMENT_SPECS:
            row = build_alignment_row(extended_header, candidate)
            included_rows.append(row)
            continue

        template, match_basis = find_template_for_candidate(candidate, uk_by_old_exact, uk_by_old_lower)
        if template:
            row = build_exact_row(extended_header, template, candidate, "Matched UK", match_basis or "UK old name")
            included_rows.append(row)
            used_uk_rows.add(template["_row_number"])
            continue

        template = choose_template_by_new(candidate.source_name, candidate.java_class, uk_by_new)
        if template:
            row = build_exact_row(extended_header, template, candidate, "Matched UK", "UK variable name")
            included_rows.append(row)
            used_uk_rows.add(template["_row_number"])
            continue

        excluded_rows.append(
            {
                "Java class": candidate.java_class,
                "Raw name": candidate.source_name,
                "Source kind": "Java field",
                "Reason": "Unmapped after review; not added to the core UK-style codebook.",
            }
        )

    included_rows = sort_codebook_rows(included_rows)
    collision_rows = build_collision_rows(included_rows)
    uk_issue_rows = build_uk_review_issues(uk_rows)

    reset_sheet(variables_ws)
    write_header(variables_ws, extended_header)
    write_rows(variables_ws, extended_header, included_rows)
    autofit_columns(variables_ws)

    update_cover_sheet(workbook)

    collision_ws = create_or_replace_sheet(workbook, "Collision Review")
    collision_header = [
        "Variable name",
        "Java class",
        "SimPaths old name",
        "EU DB column / alias",
        "Comparison status",
        "Needs review",
        "Review note",
    ]
    write_header(collision_ws, collision_header)
    for row_number, row in enumerate(collision_rows, start=2):
        for column_number, column_name in enumerate(collision_header, start=1):
            collision_ws.cell(row=row_number, column=column_number, value=row.get(column_name))
        for column_number in range(1, len(collision_header) + 1):
            collision_ws.cell(row=row_number, column=column_number).fill = REVIEW_FILL
        for column_name in ("Review note",):
            collision_ws.cell(row=row_number, column=collision_header.index(column_name) + 1).alignment = Alignment(wrap_text=True, vertical="top")
    autofit_columns(collision_ws)

    uk_issues_ws = create_or_replace_sheet(workbook, "UK Review Issues")
    uk_issue_header = ["Issue type", "Variable name", "SimPaths old name", "Java class", "Description", "Review note"]
    write_header(uk_issues_ws, uk_issue_header)
    for row_number, row in enumerate(uk_issue_rows, start=2):
        for column_number, column_name in enumerate(uk_issue_header, start=1):
            uk_issues_ws.cell(row=row_number, column=column_number, value=row.get(column_name))
        for column_number in range(1, len(uk_issue_header) + 1):
            uk_issues_ws.cell(row=row_number, column=column_number).fill = REVIEW_FILL
        for column_name in ("Description", "Review note"):
            uk_issues_ws.cell(row=row_number, column=uk_issue_header.index(column_name) + 1).alignment = Alignment(wrap_text=True, vertical="top")
    autofit_columns(uk_issues_ws)

    uk_only_ws = create_or_replace_sheet(workbook, "UK Only")
    uk_only_header = [
        "Variable name",
        "SimPaths old name",
        "Java class",
        "Module",
        "Description",
        "Reason",
    ]
    write_header(uk_only_ws, uk_only_header)
    uk_only_rows = []
    included_variable_names = {row.get("Variable name") for row in included_rows}
    for row in uk_rows:
        if row["_row_number"] in used_uk_rows:
            continue
        if (row.get("SimPaths old name"), row.get("Variable name")) in {
            ("idFather", "idMother"),
            ("ydses_p60", "yHhQuintilesC5P80"),
        }:
            continue
        if row.get("Variable name") in included_variable_names:
            continue
        uk_only_rows.append(
            {
                "Variable name": row.get("Variable name"),
                "SimPaths old name": row.get("SimPaths old name"),
                "Java class": row.get("Java class"),
                "Module": row.get("Module"),
                "Description": row.get("Description"),
                "Reason": "Present in reviewed UK workbook but no SimPathsEU core-variable match was found in this pass.",
                "Comparison status": "UK only",
            }
        )
    for row_number, row in enumerate(sorted(uk_only_rows, key=lambda item: (str(item["Module"]), str(item["Variable name"]))), start=2):
        for column_number, column_name in enumerate(uk_only_header, start=1):
            uk_only_ws.cell(row=row_number, column=column_number, value=row.get(column_name))
        for column_number in range(1, len(uk_only_header) + 1):
            uk_only_ws.cell(row=row_number, column=column_number).fill = UK_ONLY_FILL
        for column_name in ("Description", "Reason"):
            uk_only_ws.cell(row=row_number, column=uk_only_header.index(column_name) + 1).alignment = Alignment(wrap_text=True, vertical="top")
    autofit_columns(uk_only_ws)

    controls_ws = create_or_replace_sheet(workbook, "Controls")
    controls_header = [
        "Enum type",
        "Raw name",
        "Proposed name",
        "Status",
        "Needs review",
        "Description",
        "Notes",
    ]
    write_header(controls_ws, controls_header)
    control_rows = []
    for java_path, enum_name in CONTROL_ENUM_FILES.items():
        for raw_name in parse_enum_constants(java_path, enum_name):
            proposed_name, status, description, needs_review, notes = CONTROL_MANUAL[enum_name][raw_name]
            control_rows.append(
                {
                    "Enum type": enum_name,
                    "Raw name": raw_name,
                    "Proposed name": proposed_name,
                    "Status": status,
                    "Needs review": needs_review,
                    "Description": description,
                    "Notes": notes,
                }
            )
    for row_number, row in enumerate(control_rows, start=2):
        for column_number, column_name in enumerate(controls_header, start=1):
            controls_ws.cell(row=row_number, column=column_number, value=row.get(column_name))
        fill = row_fill_for_status(row["Status"])
        if fill:
            for column_number in range(1, len(controls_header) + 1):
                controls_ws.cell(row=row_number, column=column_number).fill = fill
        if row["Needs review"] == "Yes":
            controls_ws.cell(row=row_number, column=controls_header.index("Needs review") + 1).fill = REVIEW_FILL
            controls_ws.cell(row=row_number, column=controls_header.index("Notes") + 1).fill = REVIEW_FILL
    autofit_columns(controls_ws)

    excluded_ws = create_or_replace_sheet(workbook, "Excluded Internals")
    excluded_header = ["Java class", "Raw name", "Source kind", "Reason"]
    write_header(excluded_ws, excluded_header)
    for row_number, row in enumerate(sorted(excluded_rows, key=lambda item: (item["Java class"], item["Raw name"])), start=2):
        for column_number, column_name in enumerate(excluded_header, start=1):
            excluded_ws.cell(row=row_number, column=column_number, value=row.get(column_name))
        for column_number in range(1, len(excluded_header) + 1):
            excluded_ws.cell(row=row_number, column=column_number).fill = DEFERRED_FILL
    autofit_columns(excluded_ws)

    deferred_ws = create_or_replace_sheet(workbook, "Deferred Inventory")
    deferred_header = ["Source", "Enum", "Raw variable", "Reason", "Suggested next step"]
    write_header(deferred_ws, deferred_header)
    deferred_rows = []
    for java_path, enum_names in DEFERRED_ENUMS.items():
        for enum_name in enum_names:
            for raw_name in parse_enum_constants(java_path, enum_name):
                deferred_rows.append(
                    {
                        "Source": java_path.name,
                        "Enum": enum_name,
                        "Raw variable": raw_name,
                        "Reason": "Reviewed but left outside the core UK-style codebook because this enum is a regression term or validation series rather than a core state/output variable.",
                        "Suggested next step": "Handle in a second pass if you want regression/validation identifiers renamed as well.",
                    }
                )
    for row_number, row in enumerate(deferred_rows, start=2):
        for column_number, column_name in enumerate(deferred_header, start=1):
            deferred_ws.cell(row=row_number, column=column_number, value=row.get(column_name))
        for column_number in range(1, len(deferred_header) + 1):
            deferred_ws.cell(row=row_number, column=column_number).fill = DEFERRED_FILL
        deferred_ws.cell(row=row_number, column=deferred_header.index("Reason") + 1).alignment = Alignment(wrap_text=True, vertical="top")
        deferred_ws.cell(row=row_number, column=deferred_header.index("Suggested next step") + 1).alignment = Alignment(wrap_text=True, vertical="top")
    autofit_columns(deferred_ws)

    summary_ws = create_or_replace_sheet(workbook, "EU Review Summary")
    summary_header = ["Item", "Value"]
    write_header(summary_ws, summary_header)
    summary_rows = build_summary_rows(included_rows, uk_only_rows, excluded_rows, deferred_rows, control_rows, collision_rows, uk_issue_rows)
    for row_number, (item, value) in enumerate(summary_rows, start=2):
        summary_ws.cell(row=row_number, column=1, value=item)
        summary_ws.cell(row=row_number, column=2, value=value)
        if item == "Highlight legend":
            summary_ws.cell(row=row_number, column=1).font = Font(bold=True)
        summary_ws.cell(row=row_number, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    autofit_columns(summary_ws)

    if TEMP_OUTPUT_CODEBOOK.exists():
        TEMP_OUTPUT_CODEBOOK.unlink()
    workbook.save(TEMP_OUTPUT_CODEBOOK)
    TEMP_OUTPUT_CODEBOOK.replace(OUTPUT_CODEBOOK)

    print(f"Wrote {OUTPUT_CODEBOOK}")
    print(f"Core rows: {len(included_rows)}")
    print(f"UK-only rows: {len(uk_only_rows)}")
    print(f"Excluded rows: {len(excluded_rows)}")
    print(f"Deferred rows: {len(deferred_rows)}")


if __name__ == "__main__":
    main()
